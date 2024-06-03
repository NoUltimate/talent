import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment

fileNames = os.listdir('C:/Users/Administrator/PycharmProjects/talent/jixiao/list')
basic = []
all = []
infos = []
colSymbol = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
             'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF']
for fileName in fileNames:
    strList = fileName.split('.')[0].split('-')
    if len(strList) == 2:
        type = strList[0]
        name = strList[1]
    type = "企业创新类"
    # 打开Excel文件
    workbook = openpyxl.load_workbook('C:/Users/Administrator/PycharmProjects/talent/jixiao/list/' + fileName, data_only=True)
    sheet = workbook.worksheets[0]
    for (i, row1) in enumerate(sheet.iter_rows()):
        if i <= 2: continue
        talentName = sheet["B" + str(i + 1)].value
        if talentName is None or talentName == "": break
        seq = 0
        valueList1 = []
        valueList2 = []
        valueList3 = []
        for (j, symbol) in enumerate(colSymbol):
            if j > 11: continue
            if sheet[symbol + str(i + 1)].value is not None:
                value = str(sheet[symbol + str(i + 1)].value)
                valueList1.append(value)
            else:
                valueList1.append("")
        for (j, symbol) in enumerate(colSymbol):
            if j <= 11 and j != 1: continue
            if sheet[symbol + str(i + 1)].value is not None:
                value = str(sheet[symbol + str(i + 1)].value)
                valueList2.append(value)
            else:
                valueList2.append("")
        for (j, symbol) in enumerate(colSymbol):
            if sheet[symbol + str(i + 1)].value is not None:
                value = str(sheet[symbol + str(i + 1)].value)
                valueList3.append(value)
            else:
                valueList3.append("")
        basic.append(valueList1)
        all.append(valueList3)
        infos.append(valueList2)

    if type == '创业类':
        for(i, info) in enumerate(infos):
            workbook = openpyxl.load_workbook(
                'C:/Users/Administrator/PycharmProjects/talent/jixiao/mb/评价体系（3创业类）.xlsx', data_only=True)
            sheet = workbook["评价体系（3创业类）"]
            seq = 4
            for (j, v) in enumerate(info):
                if seq > 23 or j == 0: continue
                sheet["C" + str(seq)].value = v
                seq = seq + 1
            workbook.save(
                'C:/Users/Administrator/PycharmProjects/talent/jixiao/result/' + type + "-" + info[0] + ".xlsx")

    if type == '企业创新类':
        for (i, info) in enumerate(infos):
            workbook = openpyxl.load_workbook(
                'C:/Users/Administrator/PycharmProjects/talent/jixiao/mb/评价体系（2企业创新类）.xlsx', data_only=True)
            sheet = workbook["评价体系（2企业创新类）"]
            seq = 6
            for (j, v) in enumerate(info):
                if seq > 19 or j == 0: continue
                sheet["C" + str(seq)].value = v
                seq = seq + 1
            workbook.save('C:/Users/Administrator/PycharmProjects/talent/jixiao/result/' + type + "-" + info[0] + ".xlsx")

    if type == '高校院所创新类':
        for (i, info) in enumerate(infos):
            workbook = openpyxl.load_workbook(
                'C:/Users/Administrator/PycharmProjects/talent/jixiao/mb/评价体系（1-高校院所创新类）.xlsx', data_only=True)
            sheet = workbook["评价体系（1-高校院所创新类）"]
            seq = 6
            for (j, v) in enumerate(info):
                if seq > 17 or j == 0: continue
                if j == 11:
                    sheet["D15"] = v
                    continue
                sheet["C" + str(seq)].value = v
                seq = seq + 1
            workbook.save('C:/Users/Administrator/PycharmProjects/talent/jixiao/result/' + type + "-" + info[0] + ".xlsx")

workbook = openpyxl.load_workbook(
        'C:/Users/Administrator/PycharmProjects/talent/jixiao/mb/附件2-1_XX市（单位）国家、省引才计划情况统计.xlsx', data_only=True)
sheet = workbook["在岗人员有关情况统计表"]
seq = 5
for (i, v1) in enumerate(basic):
    for (j, v2) in enumerate(v1):
        sheet[colSymbol[j] + str(seq)].value = v2
    seq = seq + 1

workbook.save('C:/Users/Administrator/PycharmProjects/talent/jixiao/result/在岗人员有关情况统计表.xlsx')

if type == "创业类":
    workbook = openpyxl.load_workbook(
            'C:/Users/Administrator/PycharmProjects/talent/jixiao/mb/创业类人才绩效评估统计表.xlsx', data_only=True)
    sheet = workbook["Sheet1"]
    seq = 4
    for (i, v1) in enumerate(all):
        for (j, v2) in enumerate(v1):
            sheet[colSymbol[j] + str(seq)].value = v2
        seq = seq + 1

    workbook.save('C:/Users/Administrator/PycharmProjects/talent/jixiao/result/创业类人才绩效评估统计汇总表.xlsx')

if type == "企业创新类":
    workbook = openpyxl.load_workbook(
            'C:/Users/Administrator/PycharmProjects/talent/jixiao/mb/企业创新青年类人才绩效评估统计表.xlsx', data_only=True)
    sheet = workbook["Sheet1"]
    seq = 4
    for (i, v1) in enumerate(all):
        for (j, v2) in enumerate(v1):
            sheet[colSymbol[j] + str(seq)].value = v2
        seq = seq + 1

    workbook.save('C:/Users/Administrator/PycharmProjects/talent/jixiao/result/企业创新青年类人才绩效评估统计汇总表.xlsx')

if type == "高校院所创新类":
    workbook = openpyxl.load_workbook(
            'C:/Users/Administrator/PycharmProjects/talent/jixiao/mb/高校院所人才创新类绩效评估统计表.xlsx', data_only=True)
    sheet = workbook["Sheet1"]
    seq = 4
    for (i, v1) in enumerate(all):
        for (j, v2) in enumerate(v1):
            sheet[colSymbol[j] + str(seq)].value = v2
        seq = seq + 1

    workbook.save('C:/Users/Administrator/PycharmProjects/talent/jixiao/result/高校院所创新类人才绩效评估统计汇总表.xlsx')



