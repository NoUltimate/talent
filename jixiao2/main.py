import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment

fileNames = os.listdir('C:/Users/Administrator/PycharmProjects/talent/jixiao2/list')
infos = []
colSymbol = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
             'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF']
type = "创业类"
for fileName in fileNames:
    print(fileName)
    end = 22
    count = 4
    if type == "创业类":
        end = 22
    if type == "高校院所":
        end = 22
    workbook = openpyxl.load_workbook('C:/Users/Administrator/PycharmProjects/talent/jixiao2/list/' + fileName)
    page = 0
    sheet = workbook.worksheets[page]
    check = False
    while page < len(workbook.get_sheet_names()) - 1:
        while count < end:
            if str(sheet["A" + str(count)].value).find("其他标志性成果") != -1:
                break
            if sheet["C" + str(count)].value is not None:
                check = True
            count = count + 1
        if check is False and page < len(workbook.get_sheet_names()) - 1:
            page = page + 1
            sheet = workbook.worksheets[page]
        else:
            break
    info = [fileName.split(".")[0].split("--")[1].replace(" ", "").split("_")[0]]
    flag = False
    for (i, row1) in enumerate(sheet.iter_rows()):
        v = sheet["C" + str(i + 1)].value
        if flag:
            if v is None:
                info.append("")
            else:
                info.append(v)
        if str(v) == "数量":
            flag = True
        if str(sheet["A" + str(i + 1)].value).find("其他标志性成果") != -1:
            v2 = sheet["B" + str(i + 1)].value
            if v2 is not None and len(str(v2)) > 0:
                info[len(info) - 1] = v2
            break
    if type == "高校院所":
        info.append("")
        info[len(info)-1] = info[len(info)-2]
        info[len(info)-2] = ""
    if info[0] == '熊炜':
        print(info[0], page, count, info)
    infos.append(info)

workbook = openpyxl.load_workbook(
    'C:/Users/Administrator/PycharmProjects/talent/jixiao2/mb/企业创新青年类人才绩效评估统计表.xlsx')
if type == "创业类":
    workbook = openpyxl.load_workbook(
        'C:/Users/Administrator/PycharmProjects/talent/jixiao2/mb/创业类人才绩效评估统计表.xlsx')
if type == "高校院所":
    workbook = openpyxl.load_workbook(
        'C:/Users/Administrator/PycharmProjects/talent/jixiao2/mb/高校院所人才创新类绩效评估统计表.xlsx')
sheet = workbook.worksheets[0]

for (i, info) in enumerate(infos):
    sheet["B" + str(i + 4)] = info[0]
    seq = 12
    for (j, v) in enumerate(info):
        if j == 0: continue
        sheet[colSymbol[seq] + str(i + 4)] = v
        seq = seq + 1

saveFileName = "创新类人才绩效评估统计表.xlsx"
if type == "创业类":
    saveFileName = "创业类人才绩效评估统计表.xlsx"
if type == "高校院所":
    saveFileName = "高校院所人才创新类绩效评估统计表.xlsx"
workbook.save('C:/Users/Administrator/PycharmProjects/talent/jixiao2/result/' + saveFileName)
