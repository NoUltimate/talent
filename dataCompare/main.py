import openpyxl
from openpyxl.styles import Alignment

# 打开Excel文件
workbook = openpyxl.load_workbook('去重后人才数据.xlsx')
# 选择第一个工作表
sheet1 = workbook["Sheet1"]
companyDataMap = {}
for (i, row1) in enumerate(sheet1.iter_rows()):
    if i == 0: continue
    name = str(sheet1["C" + str(i + 1)].value)
    id = str(sheet1["F" + str(i + 1)].value)
    companyName = str(sheet1["AG" + str(i + 1)].value)
    level = str(sheet1["AK" + str(i + 1)].value)
    countMap = companyDataMap.get(companyName, {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0})
    if companyName == "None":
        continue
    countMap[level] = countMap[level] + 1
    companyDataMap[companyName] = countMap

# 打开Excel文件
workbook = openpyxl.load_workbook('全区重点生物医药企业0602.xlsx')
# 选择第一个工作表
sheet2 = workbook["杭州生物医药企业产业链匹配表"]

for (i, row1) in enumerate(sheet2.iter_rows()):
    companyName = str(sheet2["B" + str(i + 1)].value)
    companyName = companyName.replace(" ", "").replace("(", "（").replace(")", "）").replace("\n", "")
    if companyDataMap.get(companyName):
        totalCount = 0
        word = ""
        for(level, count) in companyDataMap[companyName].items():
            if count == 0:
                continue
            totalCount += count
            word += level + '类' + str(count) + "人、"
        word = "共" + str(totalCount) + "人，其中" + word.rstrip("、")
        print(companyName, word)